import numpy as np
import pandas as pd
import os
import glob

import math 
import openpyxl
import csv
import tensorflow as tf
from datetime import date
from sklearn.model_selection import train_test_split
from preprocessing import * 


class DataManagerSurvival(object):
    """A class to read the excel file with survival data, get CT/PET path, and prepare train, val and test set

    Args:
        object ([type]): [description]
    """

    def __init__(self, base_path=None,excel_path="FLIP.xlsx", csv_path=None, mask= False, mixed_data=False, columns=None):
        self.base_path = base_path
        self.excel_path=excel_path
        self.csv_path = csv_path
        self.seed = 42  # random state
        self.test_size = 0.0
        self.val_size = 0.2
        self.mixed_data=mixed_data
        self.columns= columns
        self.mask = mask

    def get_data_survival(self, create_csv, days_interval):
        '''
            retrieve the survival data from the excel file and the path of the nifti files from the base path 
            create a csv file containing all the data if create_csv == True
        '''
        y_time=np.array([])
        y_event=np.array([])
        PT_paths=[]
        CT_paths=[]
        MASK_paths=[]
        Anonym=[] #anonym name which is also the name of the directory in which there are the PT and CT scans 
        data_exist= True #boolean: data for a patient found
        struct_data = [] #data for 2nd neural network (complementary data : grade, OMS...)
        #open file containing survival information
        excel = openpyxl.load_workbook(self.excel_path)
        sheet = excel.active
        #for each patient
        for i in range(2, sheet.max_row):
            data_exist=True
            x = sheet.cell(row=i, column=5)
            if (x.value != None):
                #retrieve the path to the nifti files from the patient : folder with the Anonymisation name from the excel cell(row=i, column=2)
                path = self.base_path+str(sheet.cell(row=i, column=2).value)+'/'
                nifti_path_PT= glob.glob(os.path.join(path, '**/*_nifti_PT.nii'), recursive=True)
                nifti_path_CT= glob.glob(os.path.join(path, '**/*_nifti_CT.nii'), recursive=True)
                nifti_path_MASK= glob.glob(os.path.join(path, '**/*_nifti_mask.nii'), recursive=True)
                if not nifti_path_PT or not nifti_path_CT or (self.mask and not nifti_path_MASK) :
                    data_exist=False

                if data_exist:
                    #retrieve y_time and y_event from excel (date format (month/day/year))
                    x = x.value.split('/')
                    x = [int(i) for i in x]
                    diagnosis_date= date(x[2],x[0],x[1])
                    #if censored and there is a last check up date : retrieve date
                    if (sheet.cell(row=i, column=6).value==0 and sheet.cell(row=i, column=8).value!=None):
                        y= sheet.cell(row=i, column=8).value.split('/')
                        y = [int(j) for j in y]
                        last_checkup_date= date(y[2],y[0],y[1])
                    #if not censored and there is a relapse date: retrieve date 
                    elif (sheet.cell(row=i, column=6).value==1 and sheet.cell(row=i, column=7).value!=None):
                        y= sheet.cell(row=i, column=7).value.split('/')
                        y = [int(j) for j in y]
                        last_checkup_date= date(y[2],y[0],y[1])
                    elif (sheet.cell(row=i, column=7).value==None): 
                        data_exist=False
                        print("Missing required data...")

                if self.mixed_data and data_exist:
                    #cmpt=0
                    #struct_data.append([])
                    data=[]
                    for j in self.columns:
                        if (sheet.cell(row=i, column=j).value!=None):
                            data.append(sheet.cell(row=i, column=j).value)
                            #struct_data[cmpt].append(sheet.cell(row=i, column=j).value)
                        else: 
                            data_exist= False
                        #cmpt+=1
                
                if data_exist:
                    Anonym = np.append(Anonym,sheet.cell(row=i, column=2).value)
                    PT_paths=np.append(PT_paths, [nifti_path_PT[0]])
                    CT_paths=np.append(CT_paths, [nifti_path_CT[0]])
                    if self.mask: 
                        MASK_paths = np.append(MASK_paths, [nifti_path_MASK[0]])

                    #time is given in 30 days intervals 
                    time= int(((last_checkup_date-diagnosis_date).days)/days_interval)
                    if int(sheet.cell(row=i, column=6).value) ==0:
                        y_time=np.append(y_time, [-time])
                    else:
                        y_time=np.append(y_time, [time])
                    #y_event=np.append(y_event,[int(sheet.cell(row=i, column=6).value)])
                    if self.mixed_data:
                        struct_data.append(data)
            
        #y_event= y_event.astype(np.int32)
        y_time=y_time.astype(np.int32)
        #print(y_time)
        #voir pour truct data et mask et lire le fichier 
        if create_csv: 
            #creation of a csv file row: "ID,time,event,CT_path,PT_path"
            data_zip = list(zip(Anonym,y_time, y_event, CT_paths, PT_paths))
            with open(self.csv_path, 'w') as csv_file:
                wtr = csv.writer(csv_file, delimiter=',', lineterminator='\n')
                wtr.writerow(["ID,time,event,CT_path,PT_path"])
                for x in data_zip : wtr.writerow ([str(x[0])+','+str(x[1])+','+str(x[2])+','+str(x[3])+','+str(x[4])])

        self.time = y_time
        if self.mixed_data:
            if self.mask:
                self.nifti = list(zip(PT_paths, CT_paths, MASK_paths))
                self.struct_data = struct_data
                return list(zip(PT_paths, CT_paths, MASK_paths)), y_time, struct_data
            else: 
                self.nifti = list(zip(PT_paths, CT_paths))
                self.struct_data = struct_data
                return list(zip(PT_paths, CT_paths)), y_time, struct_data
        else: 
            if self.mask:
                self.nifti = list(zip(PT_paths, CT_paths, MASK_paths))
                return list(zip(PT_paths, CT_paths, MASK_paths)), y_time
            else: 
                self.nifti = list(zip(PT_paths, CT_paths))
                return list(zip(PT_paths, CT_paths)), y_time    
    

    @staticmethod
    def split_train_val_test_split(x, y, struct_data=None, test_size=0.0, val_size=0.2, random_state=42):
        """
            Splits x and y in two subsets 
        """
        if struct_data!=None:
            x_train, x_val, y_train, y_val, struct_data_train, struct_data_val = train_test_split(x, y, struct_data, test_size=val_size, random_state=random_state)
        else:
            x_train, x_val, y_train, y_val = train_test_split(x, y, test_size=val_size, random_state=random_state)

        if test_size!=0.:
            size = test_size/(1 - val_size)
            if struct_data!= None:
                x_train, x_test, y_train, y_test, struct_data_train, struct_data_test = train_test_split(x_train, y_train, struct_data_train, test_size=size, random_state=random_state)
                return x_train, x_val, x_test, y_train, y_val, y_test, struct_data_train, struct_data_val, struct_data_test
            else:
                x_train, x_test, y_train, y_test = train_test_split(x_train, y_train, test_size=size, random_state=random_state)
                return x_train, x_val, x_test, y_train, y_val, y_test

        else:
            if struct_data!=None:
                return x_train, x_val, y_train, y_val, struct_data_train, struct_data_val
            else:
                return x_train, x_val, y_train, y_val

    def get_images_paths_train_val(self, x_train,x_val):
        dataset_x = dict()
        dataset_x['train']=[]
        dataset_x['val']=[]

        for i in range(len(x_train)):
            if self.mask:
                dataset_x['train'].append({'pet_img':x_train[i][0], 'ct_img':x_train[i][1], 'mask_img':x_train[i][2]})
            else:
                dataset_x['train'].append({'pet_img':x_train[i][0], 'ct_img':x_train[i][1]})
        for i in range(len(x_val)):    
            if self.mask:
                dataset_x['val'].append({'pet_img':x_val[i][0], 'ct_img':x_val[i][1], 'mask_img':x_val[i][2]})
            else:
                 dataset_x['val'].append({'pet_img':x_val[i][0], 'ct_img':x_val[i][1]})
        
        return dataset_x['train'], dataset_x['val']

    def dataset_survival(self, create_csv, modalities, survival, mode, method, tval, target_size, target_spacing, target_direction):
        train_struct_data=[] #initialisation
        val_struct_data=[] #initialisation
        censored_y=[]
        not_censored_y= []
        censored_x=[]
        not_censored_x= []
        censored_struct=[]
        not_censored_struct = []
        if self.mixed_data:
            for i in range(len(self.time)):
                if self.time[i]>=0:
                    not_censored_y.append(self.time[i])
                    not_censored_x.append(self.nifti[i])
                    not_censored_struct.append(self.struct_data[i])
                else:
                    censored_y.append(self.time[i])
                    censored_x.append(self.nifti[i])
                    censored_struct.append(self.struct_data[i])
            x_train_rec, x_val_rec, y_train_rec, y_val_rec, train_struct_rec, val_struct_rec= self.split_train_val_test_split(not_censored_x, not_censored_y, not_censored_struct, test_size=0.0, val_size=0.5, random_state=42)
            x_train, x_val, y_train, y_val, train_struct  , val_struct = self.split_train_val_test_split(censored_x, censored_y, censored_struct, test_size=0.0, val_size=self.val_size, random_state=42)

            y_train = y_train[:int(len(y_train))]
            x_train = x_train[:int(len(x_train))]
            x_train = x_train+x_train_rec
            x_test = x_val+x_val_rec
            y_train = y_train+y_train_rec
            y_test = y_val+y_val_rec
            train_struct_data = train_struct+train_struct_rec
            val_struct_data = val_struct+val_struct_rec
            #x_train, x_val, y_train, y_val, train_struct_data, val_struct_data = DM.split_train_val_test_split(x, y, struct_data, test_size=0.0, val_size=val_size, random_state=42)
        else:
            x,y = get_data_survival(create_csv=create_csv) #x: images path, y: (time, event)
            x_train, x_val, y_train, y_val = self.split_train_val_test_split(x, y, test_size=0.0, val_size=val_size, random_state=42)
        
        batch_size_train = len(x_train)
        batch_size_val = len(x_val)
        
        train_images_paths_x, val_images_paths_x = self.get_images_paths_train_val(x_train,x_val)
        train_transforms = get_transform('train', modalities, mode, method, tval, target_size, target_spacing, target_direction, None, data_augmentation = True, from_pp=False, cache_pp=False, mask_survival= survival)
        val_transforms = get_transform('val', modalities, mode, method, tval, target_size, target_spacing, target_direction, None,  data_augmentation = False, from_pp=False, cache_pp=False, mask_survival= survival)

        train_X_batch=[]
        train_Y_batch = []
        val_X_batch = []
        val_Y_batch = []
        train_X_batch_struct = []
        val_X_batch_struct = []
        buffer_size= 100
        for i in range(len(train_images_paths_x)):
            print(train_images_paths_x[i])
            train=train_transforms(train_images_paths_x[i])
            train_X_batch.append(train['input'])
            train_Y_batch.append(y_train[i])
            if self.mixed_data:
                train_X_batch_struct.append(train_struct_data[i])
        for i in range(len(val_images_paths_x)):
            val=val_transforms(val_images_paths_x[i])
            val_X_batch.append(val['input'])
            val_Y_batch.append(y_val[i])
            if self.mixed_data: 
                val_X_batch_struct.append(val_struct_data[i])

        if self.mixed_data:
            return train_X_batch_struct, train_X_batch, train_Y_batch, val_X_batch_struct, val_X_batch, val_Y_batch
        else: 
            return  train_X_batch, train_Y_batch, val_X_batch, val_Y_batch